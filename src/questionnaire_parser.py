"""
src/questionnaire_parser.py
----------------------------
Responsável por ler questionários em formato Excel e extrair
as perguntas de forma estruturada.

Espera um Excel com pelo menos as colunas:
  - Coluna C (índice 2): texto da pergunta
  - Coluna D (índice 3): resposta (em branco, será preenchida)

A linha 4 é o cabeçalho — as perguntas começam na linha 5.
"""

import openpyxl
import io


def parse_questionnaire(excel_bytes: bytes, filename: str) -> list[dict]:
    """
    Lê um arquivo Excel e extrai as perguntas estruturadas.

    Parâmetros:
        excel_bytes: conteúdo do Excel em bytes
        filename: nome do arquivo (para mensagens de erro)

    Retorna:
        Lista de dicionários com:
          - row: número da linha no Excel (para reescrever depois)
          - num: número da pergunta (coluna A)
          - categoria: categoria da pergunta (coluna B)
          - pergunta: texto da pergunta (coluna C)
          - resposta: resposta atual (coluna D, geralmente vazia)
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
    except Exception as e:
        raise ValueError(f"Erro ao abrir o arquivo '{filename}': {str(e)}")

    perguntas = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
        # Ignora linhas completamente vazias
        if not any(row):
            continue

        num = str(row[0]).strip() if row[0] else ""
        categoria = str(row[1]).strip() if row[1] else ""
        pergunta = str(row[2]).strip() if row[2] else ""
        resposta = str(row[3]).strip() if row[3] else ""

        # Só processa linhas que têm uma pergunta
        if not pergunta or pergunta.lower() in ["none", "nan", ""]:
            continue

        perguntas.append({
            "row": row_idx,
            "num": num,
            "categoria": categoria,
            "pergunta": pergunta,
            "resposta": resposta,
        })

    if not perguntas:
        raise ValueError(
            f"Nenhuma pergunta encontrada em '{filename}'. "
            "Verifique se o arquivo segue o formato esperado "
            "(perguntas na coluna C, a partir da linha 5)."
        )

    return perguntas


def export_answers_to_excel(
    excel_bytes: bytes,
    answers: list[dict],
) -> bytes:
    """
    Preenche a coluna D (Resposta) do Excel original com as
    respostas geradas pelo QualiBot e retorna o arquivo atualizado.

    Parâmetros:
        excel_bytes: conteúdo original do Excel em bytes
        answers: lista de dicts com 'row' e 'resposta'

    Retorna:
        Bytes do Excel preenchido, pronto para download
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    from openpyxl.styles import Font, Alignment, PatternFill

    for item in answers:
        cell = ws.cell(row=item["row"], column=4)
        cell.value = item["resposta"]
        cell.font = Font(name="Arial", size=10, color="1A1A1A")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill("solid", start_color="E8F5E9", fgColor="E8F5E9")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
